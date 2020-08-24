# -*- coding: UTF-8 -*-
import requests
import pandas as pd
import configparser
from datetime import date
from openpyxl import Workbook, load_workbook
import os
import sys
import argparse
from random import randint
from time import sleep

API_URL = "https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date=%s&stockNo=%s"
CONFIG_PATH = 'config.ini'

class STOCK(object):
    def __init__(self):
        self.config = configparser.ConfigParser()
        self.config.read(CONFIG_PATH)

    def getStockInfo(self, section):
        items = self.config.items(section)
        name = items[0][1]
        interval = items[1][1]
        start = items[2][1]
        end = items[3][1]
        return name, interval, start, end

    def requestURLs(self):
        total_section = self.config.sections()
        urls = {}
        for section in total_section:
            name, interval, start, end = self.getStockInfo(section)
            urls[section] = {"name": name, "urls":[] }
            if interval:
                today = date.today()
                year = today.year
                month = today.month
                for i in range(0, int(interval)):
                    if i != 0:
                        month = month - 1
                    if month <= 0:
                        year = year - 1
                        month = 12
                    url = API_URL % ("%s%02d01" % (year, month), section)
                    urls[section]['urls'].append(url)
            urls[section]['urls'].sort()
        return urls

    def requestURLs4NewStock(self, name, stockNo, interval, start, end):
        urls = {}
        urls[stockNo] = {"name": name, "urls":[] }
        if interval:
                today = date.today()
                year = today.year
                month = today.month
                for i in range(0, int(interval)):
                    if i != 0:
                        month = month - 1
                    if month <= 0:
                        year = year - 1
                        month = 12
                    url = API_URL % ("%s%02d01" % (year, month), stockNo)
                    urls[stockNo]['urls'].append(url)
        urls[stockNo]['urls'].sort()
        return urls

    def excelWriter(self, rawData=None):
        #filename = date.today().strftime('%Y-%m-%d') + '.xlsx'
        filename = 'stockintaiwan.xlsx'
        cols = ['date', 'deal_share_num', 'the_price', 'start', 'highest', 'lowest', 'end', 'diff', 'deal_amount']
        stockName = rawData[0]['title'].split(' ')[2]

        if os.path.exists(filename):
            wb = load_workbook(filename)
        else:
            wb = Workbook()
        if stockName in wb.sheetnames:
            ws = wb[stockName]
            pass
        else:
            ws = wb.create_sheet(stockName)
            for i in range(0, len(cols)):
                ws.cell(row=1, column=i+1, value=cols[i])
        colDates = []
        for i in ws["A"]:
            colDates.append(i.value)
        for i in rawData:
            rowsize = ws.max_row
            for c in range(0, len(i['data'])):
                if not(i['data'][c][0] in colDates):
                    rowsize += 1
                    ws.cell(row=rowsize, column=1, value=i['data'][c][0])
                    ws.cell(row=rowsize, column=2, value=i['data'][c][1])
                    ws.cell(row=rowsize, column=3, value=i['data'][c][2])
                    ws.cell(row=rowsize, column=4, value=i['data'][c][3])
                    ws.cell(row=rowsize, column=5, value=i['data'][c][4])
                    ws.cell(row=rowsize, column=6, value=i['data'][c][5])
                    ws.cell(row=rowsize, column=7, value=i['data'][c][6])
                    ws.cell(row=rowsize, column=8, value=i['data'][c][7])
                    ws.cell(row=rowsize, column=9, value=i['data'][c][8])
        wb.save(filename)

    def crawller(self, new=False, **kwargs):
        if new:
            print(kwargs)
            stocks = self.requestURLs4NewStock(**kwargs)
        else:
            stocks = self.requestURLs()
        for i in stocks.keys():
            rawData = []
            for j in stocks[i]['urls']:
                sleep(randint(3,15))
                rawData.append(requests.get(j).json())
            self.excelWriter(rawData)
            sleep(5)

    def test(self):
        rawData = [{'stat': 'OK', 'date': '20200801', 'title': '109年08月 2412 中華電           各日成交資訊', 'fields': ['日期', '成交股數', '成交金額', '開盤價', '最高價', '最低價', '收盤價', '漲跌價差', '成交筆數'], 'data': [['109/08/03', '8,471,919', '919,957,667', '109.00', '109.50', '108.00', '108.00', '-1.50', '4,041'], ['109/08/04', '4,972,486', '539,801,474', '108.50', '109.00', '108.00', '108.50', '+0.50', '2,347'], ['109/08/05', '4,954,888', '537,813,792', '109.00', '109.50', '108.00', '108.00', '-0.50', '2,660'], ['109/08/06', '3,325,191', '361,624,857', '108.50', '109.00', '108.00', '109.00', '+1.00', '1,903'], ['109/08/07', '5,560,968', '604,768,512', '108.50', '109.00', '108.00', '109.00', ' 0.00', '2,654']], 'notes': ['符號說明:+/-/X表示漲/跌/不比價', '當日統計資訊含一般、零股、盤後定價、鉅額交易不含拍賣、標購。', 'ETF證券代號第六碼為K、M、S、C者，表示該ETF以外幣交易。']}]
        self.excelWriter(rawData)

if __name__ == "__main__":
    args = sys.argv
    if len(args) <= 1:
        STOCK().crawller()
    elif args[1] == "test":
        print("test sample")
        STOCK().test()
    elif args[1] == "help":
        print("python crawler.py name stockNo")
    else:
        STOCK().crawller(True, name=args[1], stockNo=args[2], interval=12, start=0, end=0)
